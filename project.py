# import openpyxl as xl
# wb = xl.load_workbook('transactions.xlsx')
# sheet = wb['Sheet1']
# cell = sheet['A1']
# cell = sheet.cell(1, 1)
# print(sheet.max_row)
#
# for row in range(2, sheet.max_row + 1):
#     sheet.cell(row, 3)
#     corrected_price = float(cell.value) * 0.9
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
def process_workbook(filename):
    # wb = xl.load_workbook('transactions.xlsx')
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)  # Column C
        value = float(cell.value)
        corrected_price = value * 0.9
        # print(f"Original: {value}, Corrected: {corrected_price}")
        corrected_price_cell = sheet.cell(row, 4)
    values = Reference(sheet,
                min_row=2,
                max_row=sheet.max_row,
                min_col=2,
                max_col=2 )
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    #wb.save('transactions2.xlsx')
    wb.save(filename)